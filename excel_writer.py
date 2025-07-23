# excel_writer.py

"""
Writes paper data and summary to an Excel file with multiple sheets.
Applies formatting for better readability.
"""

import pandas as pd
import os
import re
from datetime import datetime
from typing import List, Dict
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.cell.rich_text import TextBlock, CellRichText
from openpyxl.cell.text import InlineFont

OUTPUT_DIR = "output"
# Define column widths for readability
COLUMN_WIDTHS = {
    'Title': 60,
    'Journal': 25,
    'Published Date': 15,
    'Authors': 40,
    'Abstract': 80,
    'URL': 15,
}

def filter_papers_by_keyword(papers_data: List[Dict], keyword: str) -> List[Dict]:
    """
    Filters papers that contain the keyword in their title (case-insensitive).
    """
    filtered_papers = []
    for paper in papers_data:
        title = paper.get("Title", "")
        if keyword.lower() in title.lower():
            filtered_papers.append(paper.copy())
    return filtered_papers

def filter_papers_by_compound_keyword(papers_data: List[Dict], compound_keyword: str) -> List[Dict]:
    """
    Filters papers that contain ANY of the keywords connected by '+' in their title (case-insensitive).
    For example, 'Alphafold+ESMfold' returns papers with either 'Alphafold' OR 'ESMfold' in title.
    """
    # Split compound keyword by '+' and clean up whitespace
    sub_keywords = [k.strip() for k in compound_keyword.split('+') if k.strip()]
    
    if not sub_keywords:
        return []
    
    filtered_papers = []
    for paper in papers_data:
        title = paper.get("Title", "").lower()
        # Check if ANY of the sub-keywords is in the title (OR logic)
        if any(sub_keyword.lower() in title for sub_keyword in sub_keywords):
            filtered_papers.append(paper.copy())
    
    return filtered_papers

def create_rich_text_with_keyword_highlighting(text: str, keyword: str) -> CellRichText:
    """
    Creates rich text with only the keyword highlighted in red and bold.
    Handles multiple occurrences and case-insensitive matching.
    """
    if not text or not keyword or text == "N/A":
        return CellRichText(text or "")
    
    # Create fonts
    normal_font = InlineFont()
    highlight_font = InlineFont(color="FF0000", b=True)
    
    # Find all occurrences of keyword (case-insensitive)
    rich_text = CellRichText()
    text_lower = text.lower()
    keyword_lower = keyword.lower()
    
    last_end = 0
    start = 0
    
    while True:
        # Find next occurrence of keyword
        start = text_lower.find(keyword_lower, start)
        if start == -1:
            break
            
        # Add text before keyword
        if start > last_end:
            rich_text.append(TextBlock(normal_font, text[last_end:start]))
        
        # Add highlighted keyword
        end = start + len(keyword)
        rich_text.append(TextBlock(highlight_font, text[start:end]))
        
        last_end = end
        start = end
    
    # Add remaining text after last keyword
    if last_end < len(text):
        rich_text.append(TextBlock(normal_font, text[last_end:]))
    
    return rich_text

def create_rich_text_with_compound_keyword_highlighting(text: str, compound_keyword: str) -> CellRichText:
    """
    Creates rich text with all sub-keywords from compound keyword highlighted in red and bold.
    For example, 'Alphafold+ESMfold' will highlight both 'Alphafold' and 'ESMfold' in the text.
    Uses a simpler approach to avoid Excel corruption issues.
    """
    if not text or not compound_keyword or text == "N/A":
        return CellRichText(text or "")
    
    # Split compound keyword by '+' and clean up whitespace
    sub_keywords = [k.strip() for k in compound_keyword.split('+') if k.strip()]
    
    if not sub_keywords:
        return CellRichText(text)
    
    # Use the same logic as single keyword highlighting but apply to each sub-keyword sequentially
    # This approach is more stable and less prone to Excel corruption
    result_text = text
    text_lower = text.lower()
    
    # Check if any sub-keyword exists in the text
    has_matches = any(sub_keyword.lower() in text_lower for sub_keyword in sub_keywords)
    
    if not has_matches:
        return CellRichText(text)
    
    # Create fonts
    normal_font = InlineFont()
    highlight_font = InlineFont(color="FF0000", b=True)
    
    # Build a list of all matches across all keywords
    all_matches = []
    for sub_keyword in sub_keywords:
        sub_keyword_lower = sub_keyword.lower()
        start = 0
        while True:
            pos = text_lower.find(sub_keyword_lower, start)
            if pos == -1:
                break
            # Store (start_pos, end_pos, original_text_slice)
            all_matches.append((pos, pos + len(sub_keyword), text[pos:pos + len(sub_keyword)]))
            start = pos + len(sub_keyword)  # Move past this match to avoid infinite loop
    
    if not all_matches:
        return CellRichText(text)
    
    # Sort matches by start position and remove overlaps
    all_matches.sort(key=lambda x: x[0])
    filtered_matches = []
    
    for match in all_matches:
        start_pos, end_pos, match_text = match
        # Only add if it doesn't overlap with previous matches
        if not filtered_matches or start_pos >= filtered_matches[-1][1]:
            filtered_matches.append(match)
    
    # Build rich text safely
    rich_text = CellRichText()
    last_end = 0
    
    for start_pos, end_pos, match_text in filtered_matches:
        # Add text before this match
        if start_pos > last_end:
            before_text = text[last_end:start_pos]
            if before_text:  # Only add non-empty text blocks
                rich_text.append(TextBlock(normal_font, before_text))
        
        # Add highlighted match
        if match_text:  # Only add non-empty highlighted text
            rich_text.append(TextBlock(highlight_font, match_text))
        
        last_end = end_pos
    
    # Add remaining text after last match
    if last_end < len(text):
        remaining_text = text[last_end:]
        if remaining_text:  # Only add non-empty text blocks
            rich_text.append(TextBlock(normal_font, remaining_text))
    
    # If no text blocks were added, return plain text
    if len(rich_text) == 0:
        return CellRichText(text)
    
    return rich_text

def write_to_excel(papers_data: List[Dict], start_date: str, end_date: str, keywords: List[str] = None):
    """
    Writes the collected paper data and a summary to an Excel file.
    Creates additional sheets for each keyword if keywords are provided.
    """
    if not papers_data:
        print("No data to write. Excel file will not be created.")
        return

    # Ensure output directory exists
    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)

    # Prepare data for DataFrame
    df = pd.DataFrame(papers_data)
    
    # --- Create Hyperlinks ---
    # Create a display-friendly URL column before converting to formula
    df['URL_Display'] = df['URL']
    df['URL'] = df['URL'].apply(
        lambda x: f'=HYPERLINK("{x}", "Link")' if pd.notna(x) and "http" in x else ""
    )


    # --- Create Summary Data ---
    collection_period = f"{start_date.replace('/', '-')} to {end_date.replace('/', '-')}"
    total_papers = len(df)
    collection_timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    journal_counts = df["Journal"].value_counts().reset_index()
    journal_counts.columns = ["Journal Name", "Number of Papers"]

    # --- Write to Excel ---
    start_date_str = start_date.replace('/', '')[2:]
    end_date_str = end_date.replace('/', '')[2:]
    file_name = f"{start_date_str}_{end_date_str}_Papers.xlsx"
    file_path = os.path.join(OUTPUT_DIR, file_name)

    with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
        # --- Summary Sheet ---
        summary_df = pd.DataFrame({
            "Summary Metric": ["Collection Period", "Total Papers", "Collection Timestamp"],
            "Value": [collection_period, total_papers, collection_timestamp]
        })
        summary_df.to_excel(writer, sheet_name="Summary", index=False, header=False, startrow=1)
        journal_counts.to_excel(writer, sheet_name="Summary", index=False, startrow=len(summary_df) + 3)
        
        # --- Papers Sheet ---
        # Write the main data, keeping the original URL for display if needed
        df[['Title', 'Journal', 'Published Date', 'Authors', 'Abstract', 'URL']].to_excel(
            writer, sheet_name="Papers", index=False
        )

        # --- Keyword-specific Sheets ---
        if keywords:
            for keyword in keywords:
                # Check if this is a compound keyword (contains '+')
                if '+' in keyword:
                    filtered_papers = filter_papers_by_compound_keyword(papers_data, keyword)
                    sub_keywords = [k.strip() for k in keyword.split('+')]
                    keyword_description = f"'{keyword}' (OR: {', '.join(sub_keywords)})"
                else:
                    filtered_papers = filter_papers_by_keyword(papers_data, keyword)
                    keyword_description = f"'{keyword}'"
                
                if filtered_papers:
                    print(f"Found {len(filtered_papers)} papers containing {keyword_description} in title.")
                    
                    # Create DataFrame for filtered papers
                    keyword_df = pd.DataFrame(filtered_papers)
                    
                    # Create hyperlinks for URLs
                    keyword_df['URL'] = keyword_df['URL'].apply(
                        lambda x: f'=HYPERLINK("{x}", "Link")' if pd.notna(x) and "http" in x else ""
                    )
                    
                    # Write to sheet with keyword name
                    sheet_name = f"Keyword={keyword}"
                    keyword_df[['Title', 'Journal', 'Published Date', 'Authors', 'Abstract', 'URL']].to_excel(
                        writer, sheet_name=sheet_name, index=False
                    )
                else:
                    print(f"No papers found containing {keyword_description} in title.")

        # --- Apply Formatting ---
        worksheet_papers = writer.sheets["Papers"]
        worksheet_summary = writer.sheets["Summary"]

        # Set column widths and text wrapping for Papers sheet
        for i, col_name in enumerate(df[['Title', 'Journal', 'Published Date', 'Authors', 'Abstract', 'URL']].columns, 1):
            col_letter = get_column_letter(i)
            width = COLUMN_WIDTHS.get(col_name, 20) # Default width 20
            worksheet_papers.column_dimensions[col_letter].width = width
            
            # Apply text wrapping to Title and Abstract
            if col_name in ["Title", "Abstract"]:
                for cell in worksheet_papers[col_letter]:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')

        # Add AutoFilter to Papers sheet
        worksheet_papers.auto_filter.ref = worksheet_papers.dimensions

        # Set column widths for Summary sheet
        worksheet_summary.column_dimensions['A'].width = 25
        worksheet_summary.column_dimensions['B'].width = 40

        # Apply formatting to keyword sheets
        if keywords:
            for keyword in keywords:
                sheet_name = f"Keyword={keyword}"
                if sheet_name in writer.sheets:
                    worksheet_keyword = writer.sheets[sheet_name]
                    
                    # Set column widths and text wrapping for keyword sheets (same as Papers sheet)
                    for i, col_name in enumerate(['Title', 'Journal', 'Published Date', 'Authors', 'Abstract', 'URL'], 1):
                        col_letter = get_column_letter(i)
                        width = COLUMN_WIDTHS.get(col_name, 20) # Default width 20
                        worksheet_keyword.column_dimensions[col_letter].width = width
                        
                        # Apply text wrapping to Title and Abstract
                        if col_name in ["Title", "Abstract"]:
                            for cell in worksheet_keyword[col_letter]:
                                cell.alignment = Alignment(wrap_text=True, vertical='top')
                    
                    # Apply keyword highlighting to titles and abstracts
                    for i, col_name in enumerate(['Title', 'Journal', 'Published Date', 'Authors', 'Abstract', 'URL'], 1):
                        if col_name in ["Title", "Abstract"]:
                            col_letter = get_column_letter(i)
                            for row_idx, cell in enumerate(worksheet_keyword[col_letter], 1):
                                if row_idx > 1 and cell.value:  # Skip header row
                                    try:
                                        # Create rich text with appropriate highlighting function
                                        if '+' in keyword:
                                            rich_text = create_rich_text_with_compound_keyword_highlighting(str(cell.value), keyword)
                                        else:
                                            rich_text = create_rich_text_with_keyword_highlighting(str(cell.value), keyword)
                                        
                                        # Validate rich text before applying (CellRichText acts like a list)
                                        if rich_text and len(rich_text) > 0:
                                            cell.value = rich_text
                                        else:
                                            # Fallback to plain text if rich text generation failed
                                            cell.value = str(cell.value)
                                    except Exception as e:
                                        # If rich text fails, keep original text
                                        print(f"Warning: Rich text highlighting failed for cell, using plain text: {e}")
                                        cell.value = str(cell.value)
                    
                    # Add AutoFilter to keyword sheet
                    worksheet_keyword.auto_filter.ref = worksheet_keyword.dimensions


    print("\nSuccessfully created Excel file: {}".format(file_path))